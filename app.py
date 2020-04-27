import openpyxl as xl
from openpyxl.chart import BarChart, Reference

#processing thousands of spreadsheet
#automate repetitive boring stuff that waste your time
def process_excel(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    for row in range(2, sheet.max_row+1):
        cell3 = sheet.cell(row, 3) #row 3
        corrected_price = cell3.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price
    values = Reference(sheet, 
                min_row=2, 
                max_row=sheet.max_row,
                min_col=4,
                max_col=4)
    chart = BarChart()
    chart.add_data(values)
    #adding the chart to the spreadsheet,to the coordinate e2
    sheet.add_chart(chart, 'e2')

    #then save when youre done
    wb.save(filename)

